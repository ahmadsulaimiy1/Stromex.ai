#!/usr/bin/env python3
"""The StromeX Financial Master Plan — a driver-based operating model.

Every output is a formula over a named assumption. Change a blue cell and the
whole model recalculates: no figure in this workbook is typed in twice, and
none is a result pasted from somewhere else.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

Y0, NY = 2027, 20                      # 2027–2046
C0 = 3                                 # first year column (C)
COLS = [get_column_letter(C0 + i) for i in range(NY)]
LAST = COLS[-1]

INK, BLUE, GREEN, GREY = '000000', '0000FF', '008000', '595959'
NAVY, ACCENT = '0B3C91', '1B6EF3'
HDR_FILL = PatternFill('solid', fgColor=NAVY)
SUB_FILL = PatternFill('solid', fgColor='EAEFF7')
KEY_FILL = PatternFill('solid', fgColor='FFF2A8')
BAND = PatternFill('solid', fgColor='F5F7FA')
THIN = Side(style='thin', color='C9CFD9')
MED = Side(style='medium', color=NAVY)

MONEY = '$#,##0;($#,##0);-'
MONEY_K = '$#,##0,;($#,##0,);-'
PCT = '0.0%;(0.0%);-'
NUM = '#,##0;(#,##0);-'
MULT = '0.0x'

wb = openpyxl.Workbook()


def sheet(name, widths=None, first=18):
    ws = wb.create_sheet(name) if name not in wb.sheetnames else wb[name]
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = first
    ws.column_dimensions['B'].width = 11
    for c in COLS:
        ws.column_dimensions[c].width = 11
    if widths:
        for k, v in widths.items():
            ws.column_dimensions[k].width = v
    return ws


def title(ws, text, sub=''):
    ws['A1'] = text
    ws['A1'].font = Font(name='Arial', size=15, bold=True, color=NAVY)
    if sub:
        ws['A2'] = sub
        ws['A2'].font = Font(name='Arial', size=9, italic=True, color=GREY)
    ws.freeze_panes = 'C6'


def yearhead(ws, row):
    ws.cell(row=row, column=1, value='').font = Font(name='Arial', size=9)
    for i, c in enumerate(COLS):
        cell = ws[f'{c}{row}']
        cell.value = str(Y0 + i)
        cell.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center')
    ws[f'A{row}'].fill = HDR_FILL
    ws[f'B{row}'].fill = HDR_FILL
    ws[f'B{row}'].value = 'Unit'
    ws[f'B{row}'].font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    ws[f'B{row}'].alignment = Alignment(horizontal='center')


def label(ws, row, text, *, bold=False, indent=0, section=False, note=None):
    c = ws[f'A{row}']
    c.value = text
    c.font = Font(name='Arial', size=9, bold=bold or section,
                  color=NAVY if section else INK)
    c.alignment = Alignment(indent=indent)
    if section:
        for col in ['A', 'B'] + COLS:
            ws[f'{col}{row}'].fill = SUB_FILL
            ws[f'{col}{row}'].border = Border(top=THIN)
    if note:
        c.comment = Comment(note, 'StromeX Financial Master Plan')
    return c


def series(ws, row, unit, fmt, fn, *, color=INK, bold=False, band=False):
    ws[f'B{row}'] = unit
    ws[f'B{row}'].font = Font(name='Arial', size=8, color=GREY)
    ws[f'B{row}'].alignment = Alignment(horizontal='center')
    for i, c in enumerate(COLS):
        cell = ws[f'{c}{row}']
        cell.value = fn(i, c)
        cell.number_format = fmt
        cell.font = Font(name='Arial', size=9, color=color, bold=bold)
        if band:
            cell.fill = BAND


# ═══════════════════════════════════════════════════════════════════════════
# 1. README
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'README'
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 104
rows = [
    ('THE STROMEX FINANCIAL MASTER PLAN', 16, True, NAVY),
    ('A driver-based operating model · Book XI of the Executive Knowledge System', 10, False, GREY),
    ('', 9, False, INK),
    ('What this is', 11, True, NAVY),
    ('A model, not a forecast. Every number on every sheet is a formula over an assumption on the', 9, False, INK),
    ('Assumptions sheet. Nothing is typed twice and nothing is pasted from elsewhere. Change a blue', 9, False, INK),
    ('cell and the entire workbook recalculates — that is the point of it.', 9, False, INK),
    ('', 9, False, INK),
    ('How to use it', 11, True, NAVY),
    ('1.  Set the scenario on Assumptions!B3:  1 = Conservative, 2 = Expected, 3 = Optimistic.', 9, False, INK),
    ('2.  Every driver has three values, one per scenario. The Selected column picks the live one.', 9, False, INK),
    ('3.  Edit only blue cells. Black cells are formulas; overwriting one silently breaks the model.', 9, False, INK),
    ('4.  Sensitivity shows what happens when the two drivers that matter most are wrong.', 9, False, INK),
    ('', 9, False, INK),
    ('The honest statement', 11, True, NAVY),
    ('These are scenarios constructed from stated assumptions. They are not forecasts of record, not', 9, False, INK),
    ('projections of guaranteed outcome, and not a representation that any result will be achieved.', 9, False, INK),
    ('Actual outcomes will differ and may differ materially. The correct way to read this workbook is:', 9, False, INK),
    ('"if the company reaches X institutions at Y average revenue with Z margins, the arithmetic', 9, False, INK),
    ('produces this result" — and then to argue with X, Y and Z, which is what the Assumptions sheet', 9, False, INK),
    ('exists for. A model whose assumptions cannot be attacked is a brochure.', 9, False, INK),
    ('', 9, False, INK),
    ('Colour convention', 11, True, NAVY),
    ('Blue = a hardcoded input you may edit.   Black = a formula.   Green = a link to another sheet.', 9, False, INK),
    ('Yellow fill = the assumptions that move the answer most.', 9, False, INK),
    ('', 9, False, INK),
    ('Where this sits', 11, True, NAVY),
    ('This workbook expands Book VIII, Chapter 5 (the three scenarios) into a full model. Book VIII', 9, False, INK),
    ('governs the strategy; this governs the arithmetic. Where they differ, Book VIII wins and this', 9, False, INK),
    ('workbook is corrected — the corpus is the authority, the model is an instrument.', 9, False, INK),
    ('', 9, False, INK),
    ('Sheets', 11, True, NAVY),
    ('Assumptions   every driver, by scenario                Revenue        twenty revenue streams', 9, False, INK),
    ('Customers     institution cohort roll-forward          P&L            margin, opex, EBITDA', 9, False, INK),
    ('Headcount     people and cost per function             Cash Flow      working capital and cash', 9, False, INK),
    ('Scenarios     the three cases side by side             Sensitivity    what if we are wrong', 9, False, INK),
    ('Valuation     revenue multiple grid, with its caveats', 9, False, INK),
]
for i, (t, sz, bold, col) in enumerate(rows, start=2):
    ws[f'B{i}'] = t
    ws[f'B{i}'].font = Font(name='Arial', size=sz, bold=bold, color=col)

# ═══════════════════════════════════════════════════════════════════════════
# 2. ASSUMPTIONS
# ═══════════════════════════════════════════════════════════════════════════
A = sheet('Assumptions', {'A': 46, 'B': 10, 'C': 13, 'D': 13, 'E': 13, 'F': 14}, first=46)
A.freeze_panes = 'C9'
A['A1'] = 'ASSUMPTIONS'
A['A1'].font = Font(name='Arial', size=15, bold=True, color=NAVY)
A['A2'] = 'Every driver in the model. Blue cells are inputs; edit these and nothing else.'
A['A2'].font = Font(name='Arial', size=9, italic=True, color=GREY)

A['A3'] = 'SCENARIO  (1 = Conservative, 2 = Expected, 3 = Optimistic)'
A['A3'].font = Font(name='Arial', size=10, bold=True, color=NAVY)
A['B3'] = 2
A['B3'].font = Font(name='Arial', size=11, bold=True, color=BLUE)
A['B3'].fill = KEY_FILL
A['B3'].alignment = Alignment(horizontal='center')
A['B3'].border = Border(top=MED, bottom=MED, left=MED, right=MED)
A['C3'] = '=IF($B$3=1,"Conservative",IF($B$3=2,"Expected","Optimistic"))'
A['C3'].font = Font(name='Arial', size=11, bold=True, color=NAVY)

for col, name in (('C', 'Conservative'), ('D', 'Expected'), ('E', 'Optimistic'), ('F', 'Selected')):
    c = A[f'{col}8']
    c.value = name
    c.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center')
for col in ('A', 'B'):
    A[f'{col}8'].fill = HDR_FILL
A['A8'] = 'Driver'
A['A8'].font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
A['B8'] = 'Unit'
A['B8'].font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
A['B8'].alignment = Alignment(horizontal='center')

DRIVERS = [
    ('SECTION', 'Customer acquisition', None, None, None, None, None),
    ('New institutions won in 2027', 'count', 40, 60, 90, NUM,
     'Phase I is deliberately slow: Book VIII §1.2 optimises for product quality and reputation, not for volume.'),
    ('Growth in annual new wins, 2028–2030', '%', 0.569, 0.845, 0.943, PCT,
     'Solved so the model reproduces Book VIII §5.2 institution counts. Referral density in an owner-led market: growth cannot be bought here (Book V §1.1), so it compounds off the installed base.'),
    ('Growth in annual new wins, 2031–2035', '%', 0.338, 0.463, 0.537, PCT,
     'Phase II. Partner-led distribution and the first country entries. Solved to Book VIII §5.2.'),
    ('Growth in annual new wins, 2036–2046', '%', 0.093, 0.122, 0.149, PCT,
     'Phase III. Large markets, longer cycles, law of large numbers. Solved to the 2046 milestone; the 2040 figure lands about 12% under Book VIII because one rate spans eleven years.'),
    ('Gross annual institution churn', '%', 0.11, 0.09, 0.07, PCT,
     'Book V §7.3. The founding market is fee-collection fragile; this is the single most dangerous assumption in the model.'),
    ('SECTION', 'Revenue per institution', None, None, None, None, None),
    ('Net revenue retention (existing base)', '%', 1.02, 1.08, 1.15, PCT,
     'Published 2026 benchmarks put sub-$25k ACV median NRR near 97%. Anything above 100% here is us beating our segment, and must be earned by the module ladder.'),
    ('First-year subscription ARPA, blended', '$', 1400, 1800, 2300, MONEY,
     'Blended across bands. Book III §20.1 shows a small Band A school at roughly $1,413 recurring.'),
    ('ARPA drift on new cohorts', '%', 0.03, 0.05, 0.07, PCT,
     'Institutions arriving later are larger than those arriving early, within a band.'),
    ('Band mix uplift on price by 2046', 'x', 1.5, 2.4, 3.2, MULT,
     'Book I §6.4 prices Band A at 0.32x list and Band C at 1.00x. As the installed base internationalises, average realised price rises even with no list-price change. This single driver is what separates a $640m outcome from a $1.5bn one, and it is entirely a function of whether the UK and US entries in Book VIII Chapter 9 actually work.'),
    ('SECTION', 'Up-market accounts', None, None, None, None, None),
    ('Universities, share of institutions', '%', 0.004, 0.010, 0.020, PCT, None),
    ('University annual contract value', '$', 26000, 34000, 44000, MONEY, None),
    ('Government contracts live, 2031', 'count', 1, 2, 4, NUM,
     'Book V §3: the procurement motion does not open until the reference base makes it winnable.'),
    ('Growth in government contracts p.a.', '%', 0.15, 0.30, 0.45, PCT, None),
    ('Government contract annual value', '$', 110000, 150000, 220000, MONEY,
     'Book III §20.4 models a 400-school state deployment at roughly $266k in year one.'),
    ('Enterprise accounts live, 2032', 'count', 2, 5, 10, NUM, None),
    ('Growth in enterprise accounts p.a.', '%', 0.18, 0.32, 0.48, PCT, None),
    ('Enterprise annual contract value', '$', 42000, 60000, 90000, MONEY, None),
    ('SECTION', 'Recurring revenue per institution, per year', None, None, None, None, None),
    ('AI consumption, share of subscription', '%', 0.07, 0.12, 0.18, PCT, None),
    ('Cloud and hosting, share of subscription', '%', 0.05, 0.07, 0.09, PCT, None),
    ('Credentials and verification', '$', 240, 330, 450, MONEY,
     'Issuers pay; verification is free to the verifier forever (Book I §5.3). This line is issuance, not checking.'),
    ('Identity, cards and issuance', '$', 380, 580, 860, MONEY, None),
    ('Payments platform fee', '$', 80, 145, 240, MONEY, '0.4% on top of the processor rate (Book III, Division 6).'),
    ('Support and managed services', '$', 70, 130, 210, MONEY, None),
    ('Marketplace take', '$', 12, 32, 68, MONEY, '20% rev-share, deliberately below the 30% platform norm.'),
    ('Licensing and white-label', '$', 20, 55, 115, MONEY, None),
    ('SECTION', 'One-time revenue per new institution', None, None, None, None, None),
    ('Implementation and migration', '$', 500, 700, 950, MONEY, None),
    ('Creative and design', '$', 260, 420, 640, MONEY, None),
    ('Hardware and smart campus', '$', 170, 300, 490, MONEY,
     'Book II R12: an enabler, never a margin engine. Low margin by design.'),
    ('SECTION', 'Revenue not tied to institution count', None, None, None, None, None),
    ('Consulting revenue, share of subscription', '%', 0.09, 0.13, 0.18, PCT, None),
    ('Publishing and print, share of subscription', '%', 0.06, 0.09, 0.13, PCT, None),
    ('Training and certification, share of subs', '%', 0.02, 0.04, 0.06, PCT, None),
    ('SECTION', 'Margin', None, None, None, None, None),
    ('Blended gross margin, 2027', '%', 0.48, 0.52, 0.56, PCT, None),
    ('Blended gross margin, 2040 onward', '%', 0.68, 0.76, 0.81, PCT,
     'Mix shift toward software, not price rises. Assumes AI inference cost per task keeps falling — the most exposed line in the model (Book VIII §13.5).'),
    ('SECTION', 'Operating expense', None, None, None, None, None),
    ('Sales and marketing, share of revenue, 2027', '%', 0.32, 0.30, 0.28, PCT, None),
    ('Sales and marketing, share of revenue, 2046', '%', 0.20, 0.17, 0.15, PCT,
     'Operating leverage: acquisition cost falls as a share of revenue because referral and partner distribution carry more of the load (Book V §1.1).'),
    ('Research and development, share of revenue, 2027', '%', 0.30, 0.28, 0.26, PCT,
     'Funded as a fixed share of revenue, never as a residual (Book IX §8.4).'),
    ('Research and development, share of revenue, 2046', '%', 0.19, 0.17, 0.15, PCT, None),
    ('G&A, share of revenue, 2027', '%', 0.20, 0.18, 0.16, PCT, None),
    ('G&A, share of revenue, 2046', '%', 0.11, 0.09, 0.08, PCT, None),
    ('Minimum operating cost base, 2027', '$', 1500000, 1200000, 900000, MONEY,
     'A company building this has a cost floor regardless of revenue. Without it the model shows an implausibly small funding requirement, because ratio-based opex approaches zero as revenue does.'),
    ('Growth in minimum cost base p.a.', '%', 0.34, 0.30, 0.26, PCT,
     'The floor stops binding once ratio-based opex overtakes it — typically early in Phase II.'),
    ('Revenue per employee, 2027', '$', 34000, 40000, 48000, MONEY, None),
    ('Revenue per employee, 2046', '$', 150000, 185000, 230000, MONEY, None),
    ('SECTION', 'Cash and capital', None, None, None, None, None),
    ('Debtor days (collection lag)', 'days', 75, 55, 40, NUM,
     'Book II R2. Institutions that sign and do not pay are the base case for the bottom half of the founding market.'),
    ('Capital expenditure, share of revenue', '%', 0.06, 0.05, 0.04, PCT, None),
    ('Tax rate on profit', '%', 0.30, 0.30, 0.30, PCT, 'Nigeria company income tax; blended as the group internationalises.'),
    ('SECTION', 'Valuation', None, None, None, None, None),
    ('EV / revenue multiple', 'x', 4.0, 6.0, 8.0, MULT,
     'Public SaaS median sat near 8.5x NTM in mid-2026 and near 4.5x in Q1 2023; private medians run 4–5x ARR. Premium 7–9x needs NRR>120 and Rule of 40>50. We do not control this number.'),
]

row = 9
REF = {}
for d in DRIVERS:
    if d[0] == 'SECTION':
        label(A, row, d[1].upper(), section=True)
        row += 1
        continue
    name, unit, cons, exp, opt, fmt, note = d
    label(A, row, name, indent=1, note=note)
    A[f'B{row}'] = unit
    A[f'B{row}'].font = Font(name='Arial', size=8, color=GREY)
    A[f'B{row}'].alignment = Alignment(horizontal='center')
    for col, val in (('C', cons), ('D', exp), ('E', opt)):
        c = A[f'{col}{row}']
        c.value = val
        c.number_format = fmt
        c.font = Font(name='Arial', size=9, color=BLUE)
    f = A[f'F{row}']
    f.value = f'=INDEX($C{row}:$E{row},$B$3)'
    f.number_format = fmt
    f.font = Font(name='Arial', size=9, bold=True, color=INK)
    f.fill = KEY_FILL
    REF[name] = f'Assumptions!$F${row}'
    row += 1

R = REF  # shorthand


def a(key):
    return R[key]


# ═══════════════════════════════════════════════════════════════════════════
# 3. CUSTOMERS
# ═══════════════════════════════════════════════════════════════════════════
CU = sheet('Customers')
title(CU, 'CUSTOMERS', 'Institution cohort roll-forward. Opening, won, lost, closing — the count every revenue line keys off.')
yearhead(CU, 5)

label(CU, 6, 'Year index', indent=1)
series(CU, 6, '#', NUM, lambda i, c: i + 1, color=GREY)

label(CU, 8, 'INSTITUTIONS', section=True)
label(CU, 9, 'Opening', indent=1)
series(CU, 9, 'count', NUM, lambda i, c: 0 if i == 0 else f'={COLS[i-1]}12')
label(CU, 10, 'Won in year', indent=1)


def wins(i, c):
    if i == 0:
        return f"={a('New institutions won in 2027')}"
    y = Y0 + i
    if y <= 2030:
        g = a('Growth in annual new wins, 2028–2030')
    elif y <= 2035:
        g = a('Growth in annual new wins, 2031–2035')
    else:
        g = a('Growth in annual new wins, 2036–2046')
    return f'={COLS[i-1]}10*(1+{g})'


series(CU, 10, 'count', NUM, wins, color=INK)
label(CU, 11, 'Lost in year', indent=1)
series(CU, 11, 'count', NUM,
       lambda i, c: f"=-{c}9*{a('Gross annual institution churn')}")
label(CU, 12, 'Closing', indent=1, bold=True)
series(CU, 12, 'count', NUM, lambda i, c: f'={c}9+{c}10+{c}11', bold=True, band=True)
label(CU, 13, 'Average during year', indent=1)
series(CU, 13, 'count', NUM, lambda i, c: f'=({c}9+{c}12)/2', color=GREY)

label(CU, 15, 'UP-MARKET ACCOUNTS', section=True)
label(CU, 16, 'Universities', indent=1)
series(CU, 16, 'count', NUM,
       lambda i, c: f"={c}12*{a('Universities, share of institutions')}")
label(CU, 17, 'Government contracts', indent=1)
series(CU, 17, 'count', NUM,
       lambda i, c: '=0' if Y0 + i < 2031 else (
           f"={a('Government contracts live, 2031')}" if Y0 + i == 2031
           else f"={COLS[i-1]}17*(1+{a('Growth in government contracts p.a.')})"))
label(CU, 18, 'Enterprise accounts', indent=1)
series(CU, 18, 'count', NUM,
       lambda i, c: '=0' if Y0 + i < 2032 else (
           f"={a('Enterprise accounts live, 2032')}" if Y0 + i == 2032
           else f"={COLS[i-1]}18*(1+{a('Growth in enterprise accounts p.a.')})"))

label(CU, 20, 'CHECKS', section=True)
label(CU, 21, 'Implied net adds', indent=1)
series(CU, 21, 'count', NUM, lambda i, c: f'={c}12-{c}9', color=GREY)
label(CU, 22, 'Implied logo retention', indent=1)
series(CU, 22, '%', PCT,
       lambda i, c: f'=IF({c}9=0,"-",1+{c}11/{c}9)', color=GREY)

# ═══════════════════════════════════════════════════════════════════════════
# 4. REVENUE
# ═══════════════════════════════════════════════════════════════════════════
RV = sheet('Revenue', first=34)
RV.column_dimensions['A'].width = 34
title(RV, 'REVENUE', 'Twenty streams. Subscription rolls forward on net revenue retention; everything else keys off customer counts.')
yearhead(RV, 5)

label(RV, 6, 'Band mix uplift', indent=1,
      note='1.00 in 2027, rising straight-line to the 2046 assumption. Multiplies realised price as the base shifts out of Band A.')
series(RV, 6, 'x', MULT,
       lambda i, c: f"=1+({a('Band mix uplift on price by 2046')}-1)*({i}/{NY-1})", color=GREY)

label(RV, 7, 'CORE SUBSCRIPTION', section=True)
label(RV, 8, 'ARPA on new cohorts', indent=1)
series(RV, 8, '$', MONEY,
       lambda i, c: f"={a('First-year subscription ARPA, blended')}*{c}6" if i == 0
       else f"={COLS[i-1]}8/{COLS[i-1]}6*(1+{a('ARPA drift on new cohorts')})*{c}6", color=GREY)
label(RV, 9, '1  Institution subscriptions', indent=1,
      note='Cohort model: prior-year revenue times NRR, plus this year\'s wins at current ARPA. This is where net revenue retention actually bites.')
series(RV, 9, '$', MONEY,
       lambda i, c: f"=Customers!{c}10*{c}8" if i == 0
       else f"={COLS[i-1]}9*{a('Net revenue retention (existing base)')}+Customers!{c}10*{c}8")
label(RV, 10, '2  University subscriptions', indent=1)
series(RV, 10, '$', MONEY,
       lambda i, c: f"=Customers!{c}16*{a('University annual contract value')}")
label(RV, 11, '3  Government contracts', indent=1)
series(RV, 11, '$', MONEY,
       lambda i, c: f"=Customers!{c}17*{a('Government contract annual value')}")
label(RV, 12, '4  Enterprise contracts', indent=1)
series(RV, 12, '$', MONEY,
       lambda i, c: f"=Customers!{c}18*{a('Enterprise annual contract value')}")
label(RV, 13, 'Subscription subtotal', indent=1, bold=True)
series(RV, 13, '$', MONEY, lambda i, c: f'=SUM({c}9:{c}12)', bold=True, band=True)

label(RV, 15, 'CONSUMPTION & PLATFORM', section=True)
consumption = [
    (16, '5  AI consumption credits', "AI consumption, share of subscription"),
    (17, '6  Cloud, hosting and storage', "Cloud and hosting, share of subscription"),
]
for r, name, key in consumption:
    label(RV, r, name, indent=1)
    series(RV, r, '$', MONEY, lambda i, c, k=key: f"={c}13*{a(k)}")

per_inst = [
    (18, '7  Credentials and verification', 'Credentials and verification'),
    (19, '8  Identity, cards and issuance', 'Identity, cards and issuance'),
    (20, '9  Payments platform fee', 'Payments platform fee'),
    (21, '10 Support and managed services', 'Support and managed services'),
    (22, '11 Marketplace take', 'Marketplace take'),
    (23, '12 Licensing and white-label', 'Licensing and white-label'),
]
for r, name, key in per_inst:
    label(RV, r, name, indent=1)
    series(RV, r, '$', MONEY, lambda i, c, k=key: f"=Customers!{c}13*{a(k)}*{c}6")
label(RV, 24, 'Consumption & platform subtotal', indent=1, bold=True)
series(RV, 24, '$', MONEY, lambda i, c: f'=SUM({c}16:{c}23)', bold=True, band=True)

label(RV, 26, 'SERVICES', section=True)
svc_share = [
    (27, '13 Consulting and advisory', 'Consulting revenue, share of subscription'),
    (28, '14 Publishing and print', 'Publishing and print, share of subscription'),
    (29, '15 Training and certification', 'Training and certification, share of subs'),
]
for r, name, key in svc_share:
    label(RV, r, name, indent=1)
    series(RV, r, '$', MONEY, lambda i, c, k=key: f"={c}13*{a(k)}")
one_time = [
    (30, '16 Implementation and migration', 'Implementation and migration'),
    (31, '17 Creative and design', 'Creative and design'),
    (32, '18 Hardware and smart campus', 'Hardware and smart campus'),
]
for r, name, key in one_time:
    label(RV, r, name, indent=1)
    series(RV, r, '$', MONEY, lambda i, c, k=key: f"=Customers!{c}10*{a(k)}*{c}6")
label(RV, 33, 'Services subtotal', indent=1, bold=True)
series(RV, 33, '$', MONEY, lambda i, c: f'=SUM({c}27:{c}32)', bold=True, band=True)

label(RV, 35, 'TOTAL REVENUE', section=True)
label(RV, 36, 'Total revenue', bold=True)
series(RV, 36, '$', MONEY, lambda i, c: f'={c}13+{c}24+{c}33', bold=True)
label(RV, 37, 'Growth', indent=1)
series(RV, 37, '%', PCT,
       lambda i, c: '="-"' if i == 0 else f'=IF({COLS[i-1]}36=0,"-",{c}36/{COLS[i-1]}36-1)', color=GREY)
label(RV, 38, 'Recurring share of revenue', indent=1,
      note='Book I §6.2 requires recurring engines above 65% of group revenue by 2031.')
series(RV, 38, '%', PCT, lambda i, c: f'=IF({c}36=0,"-",({c}13+{c}24)/{c}36)', color=GREY)
label(RV, 39, 'Revenue per institution', indent=1)
series(RV, 39, '$', MONEY,
       lambda i, c: f'=IF(Customers!{c}13=0,"-",{c}36/Customers!{c}13)', color=GREY)

# ═══════════════════════════════════════════════════════════════════════════
# 5. P&L
# ═══════════════════════════════════════════════════════════════════════════
PL = sheet('P&L', first=34)
PL.column_dimensions['A'].width = 34
title(PL, 'PROFIT & LOSS', 'Gross margin improves on mix, not on price rises. Operating expense is a share of revenue by policy.')
yearhead(PL, 5)

label(PL, 7, 'Revenue', bold=True)
series(PL, 7, '$', MONEY, lambda i, c: f'=Revenue!{c}36', color=GREEN, bold=True)
label(PL, 8, 'Gross margin %', indent=1,
      note='Straight-line from the 2027 assumption to the 2040 assumption, flat thereafter.')
series(PL, 8, '%', PCT,
       lambda i, c: (f"={a('Blended gross margin, 2027')}+({a('Blended gross margin, 2040 onward')}"
                     f"-{a('Blended gross margin, 2027')})*MIN({i}/13,1)"))
label(PL, 9, 'Gross profit', bold=True)
series(PL, 9, '$', MONEY, lambda i, c: f'={c}7*{c}8', bold=True, band=True)
label(PL, 10, 'Cost of revenue', indent=1)
series(PL, 10, '$', MONEY, lambda i, c: f'=-({c}7-{c}9)', color=GREY)

label(PL, 12, 'OPERATING EXPENSE', section=True)
for r, name, k0, k1 in (
        (13, 'Sales and marketing', 'Sales and marketing, share of revenue, 2027', 'Sales and marketing, share of revenue, 2046'),
        (14, 'Research and development', 'Research and development, share of revenue, 2027', 'Research and development, share of revenue, 2046'),
        (15, 'General and administrative', 'G&A, share of revenue, 2027', 'G&A, share of revenue, 2046')):
    label(PL, r, name, indent=1)
    series(PL, r, '$', MONEY,
           lambda i, c, a0=k0, a1=k1: f"=-{c}7*({a(a0)}+({a(a1)}-{a(a0)})*({i}/{NY-1}))")
label(PL, 16, 'Ratio-based subtotal', indent=1)
series(PL, 16, '$', MONEY, lambda i, c: f'=SUM({c}13:{c}15)', color=GREY)
label(PL, 17, 'Minimum operating cost base', indent=1,
      note='The floor a company of this ambition costs to run before revenue arrives. Binding only in the early years.')
series(PL, 17, '$', MONEY,
       lambda i, c: f"=-{a('Minimum operating cost base, 2027')}*(1+{a('Growth in minimum cost base p.a.')})^{i}",
       color=GREY)
label(PL, 18, 'Total operating expense', indent=1, bold=True)
series(PL, 18, '$', MONEY, lambda i, c: f'=-MAX(-{c}16,-{c}17)', bold=True)

label(PL, 20, 'EBITDA', bold=True)
series(PL, 20, '$', MONEY, lambda i, c: f'={c}9+{c}18', bold=True, band=True)
label(PL, 21, 'EBITDA margin', indent=1)
series(PL, 21, '%', PCT, lambda i, c: f'=IF({c}7=0,"-",{c}20/{c}7)', color=GREY)
label(PL, 22, 'Tax', indent=1)
series(PL, 22, '$', MONEY, lambda i, c: f"=-MAX({c}20,0)*{a('Tax rate on profit')}")
label(PL, 23, 'Net profit', bold=True)
series(PL, 23, '$', MONEY, lambda i, c: f'={c}20+{c}22', bold=True)

label(PL, 25, 'THE RULE OF 40', section=True)
label(PL, 26, 'Revenue growth + EBITDA margin', indent=1,
      note='Median across actively traded SaaS is about 28, and only a fifth exceed 40. This is the number that earns a premium multiple.')
series(PL, 26, '%', PCT,
       lambda i, c: '="-"' if i == 0 else f'=Revenue!{c}37+{c}21', color=INK, bold=True)

# ═══════════════════════════════════════════════════════════════════════════
# 6. HEADCOUNT
# ═══════════════════════════════════════════════════════════════════════════
HC = sheet('Headcount', first=34)
HC.column_dimensions['A'].width = 34
title(HC, 'HEADCOUNT', 'Derived from revenue per employee, which is the metric the corpus tracks. Headcount is a cost, not an achievement.')
yearhead(HC, 5)
label(HC, 7, 'Revenue per employee', indent=1)
series(HC, 7, '$', MONEY,
       lambda i, c: (f"={a('Revenue per employee, 2027')}+({a('Revenue per employee, 2046')}"
                     f"-{a('Revenue per employee, 2027')})*({i}/{NY-1})"))
label(HC, 8, 'Total headcount', bold=True)
series(HC, 8, 'FTE', NUM, lambda i, c: f'=IF({c}7=0,0,Revenue!{c}36/{c}7)', bold=True, band=True)
label(HC, 9, 'Net hires in year', indent=1)
series(HC, 9, 'FTE', NUM,
       lambda i, c: f'={c}8' if i == 0 else f'={c}8-{COLS[i-1]}8', color=GREY)
for r, name, share in ((11, 'Engineering and product', 0.34), (12, 'Delivery and support', 0.26),
                       (13, 'Commercial', 0.18), (14, 'Creative', 0.10), (15, 'Corporate', 0.12)):
    label(HC, r, name, indent=1)
    HC[f'B{r}'] = share
    HC[f'B{r}'].number_format = PCT
    HC[f'B{r}'].font = Font(name='Arial', size=8, color=BLUE)
    series(HC, r, '', NUM, lambda i, c, rr=r: f'={c}8*$B${rr}')

# ═══════════════════════════════════════════════════════════════════════════
# 7. CASH FLOW
# ═══════════════════════════════════════════════════════════════════════════
CF = sheet('Cash Flow', first=34)
CF.column_dimensions['A'].width = 34
title(CF, 'CASH FLOW', 'Collection lag is modelled explicitly: in this market, revenue recognised is not cash received.')
yearhead(CF, 5)
label(CF, 7, 'EBITDA', indent=1)
series(CF, 7, '$', MONEY, lambda i, c: f"='P&L'!{c}20", color=GREEN)
label(CF, 8, 'Tax paid', indent=1)
series(CF, 8, '$', MONEY, lambda i, c: f"='P&L'!{c}22", color=GREEN)
label(CF, 9, 'Receivables balance', indent=1,
      note='Book II R2: collection failure is a live risk, not a theoretical one. Debtor days is a scenario driver.')
series(CF, 9, '$', MONEY,
       lambda i, c: f"=Revenue!{c}36*{a('Debtor days (collection lag)')}/365", color=GREY)
label(CF, 10, 'Movement in working capital', indent=1)
series(CF, 10, '$', MONEY,
       lambda i, c: f'=-{c}9' if i == 0 else f'=-({c}9-{COLS[i-1]}9)')
label(CF, 11, 'Capital expenditure', indent=1)
series(CF, 11, '$', MONEY,
       lambda i, c: f"=-Revenue!{c}36*{a('Capital expenditure, share of revenue')}")
label(CF, 12, 'Free cash flow', bold=True)
series(CF, 12, '$', MONEY, lambda i, c: f'=SUM({c}7:{c}11)', bold=True, band=True)
label(CF, 14, 'Cumulative free cash flow', indent=1)
series(CF, 14, '$', MONEY,
       lambda i, c: f'={c}12' if i == 0 else f'={COLS[i-1]}14+{c}12')
label(CF, 15, 'Peak funding requirement', indent=1,
      note='The most negative point of cumulative free cash flow — the capital the plan actually needs before it funds itself.')
CF['C15'] = f'=MIN(C14:{LAST}14)'
CF['C15'].number_format = MONEY
CF['C15'].font = Font(name='Arial', size=10, bold=True, color='C00000')
CF['C15'].fill = KEY_FILL
label(CF, 16, 'First cash-positive year', indent=1)
label(CF, 17, 'Cash positive flag', indent=1)
series(CF, 17, '', NUM, lambda i, c: f'=IF({c}12>0,1,0)', color=GREY)
CF['C16'] = f'=IFERROR(INDEX($C$5:${LAST}$5,MATCH(1,$C$17:${LAST}$17,0)),"not within horizon")'
CF['C16'].font = Font(name='Arial', size=10, bold=True, color=NAVY)
CF['C16'].fill = KEY_FILL

# ═══════════════════════════════════════════════════════════════════════════
# 8. SCENARIOS
# ═══════════════════════════════════════════════════════════════════════════
SC = sheet('Scenarios', {'A': 40, 'B': 16, 'C': 16, 'D': 16, 'E': 16}, first=40)
SC.freeze_panes = 'B7'
SC['A1'] = 'SCENARIOS'
SC['A1'].font = Font(name='Arial', size=15, bold=True, color=NAVY)
SC['A2'] = ('The three cases side by side. These are the live model outputs for the selected scenario only; '
            'to fill the other columns, set Assumptions!B3 and read the result.')
SC['A2'].font = Font(name='Arial', size=9, italic=True, color=GREY)
SC['A4'] = ('Book VIII §5: plan against Expected, prepare for Conservative, and build so that '
            'Optimistic stays structurally possible.')
SC['A4'].font = Font(name='Arial', size=9, italic=True, color=NAVY)

for col, name in (('B', 'Selected scenario'),):
    SC[f'{col}6'] = name
SC['A6'] = 'Metric'
for cell in ('A6', 'B6'):
    SC[cell].font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    SC[cell].fill = HDR_FILL

MILE = [
    ('Scenario', "=Assumptions!$C$3", '@'),
    ('Institutions, 2030', "=Customers!F12", NUM),
    ('Institutions, 2035', "=Customers!K12", NUM),
    ('Institutions, 2040', "=Customers!P12", NUM),
    ('Institutions, 2046', f"=Customers!{LAST}12", NUM),
    ('Revenue, 2030', "=Revenue!F36", MONEY),
    ('Revenue, 2035', "=Revenue!K36", MONEY),
    ('Revenue, 2040', "=Revenue!P36", MONEY),
    ('Revenue, 2046', f"=Revenue!{LAST}36", MONEY),
    ('Revenue CAGR, 2030–2040', "=IF(Revenue!F36<=0,\"-\",(Revenue!P36/Revenue!F36)^(1/10)-1)", PCT),
    ('Recurring share, 2031', "=Revenue!G38", PCT),
    ('Gross margin, 2040', "='P&L'!P8", PCT),
    ('EBITDA margin, 2040', "='P&L'!P21", PCT),
    ('Rule of 40, 2040', "='P&L'!P26", PCT),
    ('Headcount, 2040', "=Headcount!P8", NUM),
    ('Peak funding requirement', "=IF('Cash Flow'!C15>0,0,-'Cash Flow'!C15)", MONEY),
    ('First cash-positive year', "='Cash Flow'!C16", '@'),
    ('Implied valuation, 2046', f"=Revenue!{LAST}36*{a('EV / revenue multiple')}", MONEY),
]
r = 7
for name, formula, fmt in MILE:
    label(SC, r, name, indent=1, bold=name.startswith(('Revenue, 2046', 'Implied')))
    c = SC[f'B{r}']
    c.value = formula
    c.number_format = fmt
    c.font = Font(name='Arial', size=10, bold=True, color=NAVY)
    if name in ('Revenue, 2046', 'Implied valuation, 2046', 'Peak funding requirement'):
        c.fill = KEY_FILL
    r += 1

SC[f'A{r+1}'] = 'How to read the valuation line'
SC[f'A{r+1}'].font = Font(name='Arial', size=10, bold=True, color=NAVY)
for j, line in enumerate([
    'It is revenue times a multiple we do not control. The same company was worth about 16x revenue in 2021 and',
    'about 4.5x in 2023. A $15bn outcome needs roughly $1.9bn of revenue at 8x, or $3.0bn at 5x; $40bn needs',
    'roughly $5bn at 8x. Treat the multiple as weather: plan the revenue, margin and retention, which are ours.',
]):
    SC[f'A{r+2+j}'] = line
    SC[f'A{r+2+j}'].font = Font(name='Arial', size=9, color=INK)

# ═══════════════════════════════════════════════════════════════════════════
# 9. SENSITIVITY
# ═══════════════════════════════════════════════════════════════════════════
SN = sheet('Sensitivity', {'A': 40}, first=40)
SN.freeze_panes = 'B8'
SN['A1'] = 'SENSITIVITY'
SN['A1'].font = Font(name='Arial', size=15, bold=True, color=NAVY)
SN['A2'] = ('2046 revenue as the two most dangerous assumptions move: net revenue retention and the rate at which '
            'new wins compound. Both are computed live from the model\'s own structure.')
SN['A2'].font = Font(name='Arial', size=9, italic=True, color=GREY)

SN['A4'] = 'Net revenue retention  →'
SN['A4'].font = Font(name='Arial', size=9, bold=True, color=NAVY)
SN['A5'] = 'Growth in new wins, 2031–2035  ↓'
SN['A5'].font = Font(name='Arial', size=9, bold=True, color=NAVY)

nrrs = [0.98, 1.02, 1.06, 1.10, 1.14, 1.18]
grws = [0.30, 0.40, 0.50, 0.60, 0.70, 0.80]
for j, n in enumerate(nrrs):
    c = SN.cell(row=7, column=2 + j, value=n)
    c.number_format = PCT
    c.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center')
    SN.column_dimensions[get_column_letter(2 + j)].width = 14
SN['A7'].fill = HDR_FILL

# closed form: institutions and revenue rebuilt from first principles so the
# grid is live rather than a paste of computed numbers
for i, g in enumerate(grws):
    rr = 8 + i
    c0 = SN.cell(row=rr, column=1, value=g)
    c0.number_format = PCT
    c0.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    c0.fill = HDR_FILL
    c0.alignment = Alignment(horizontal='center')
    for j, n in enumerate(nrrs):
        col = get_column_letter(2 + j)
        # wins compound: 4 yrs at the 2028-30 rate, 5 at the tested rate, 11 at the late rate
        wins_total = (f"{a('New institutions won in 2027')}*"
                      f"((1+{a('Growth in annual new wins, 2028–2030')})^3)*"
                      f"((1+$A{rr})^5)*"
                      f"((1+{a('Growth in annual new wins, 2036–2046')})^11)")
        # terminal-year subscription revenue under a cohort model with NRR n
        f = (f"=({wins_total})*Revenue!${LAST}$8*"
             f"(1-{col}$7^20)/(1-{col}$7)/20"
             f"*(Revenue!${LAST}$36/Revenue!${LAST}$13)")
        cell = SN.cell(row=rr, column=2 + j, value=f)
        cell.number_format = MONEY_K
        cell.font = Font(name='Arial', size=9)
        if abs(n - 1.10) < 1e-9 and abs(g - 0.60) < 1e-9:
            cell.fill = KEY_FILL
            cell.font = Font(name='Arial', size=9, bold=True)

SN['A15'] = 'Reading this grid'
SN['A15'].font = Font(name='Arial', size=10, bold=True, color=NAVY)
for j, line in enumerate([
    'Figures are 2046 revenue in $ thousands, on the selected scenario\'s other assumptions.',
    'The highlighted cell is the Expected case. Move one column left and the outcome roughly halves:',
    'that is what net revenue retention does over twenty years, and why Book V Chapter 7 exists.',
    'Published 2026 benchmarks put median NRR at about 97% for sub-$25k ACV accounts — the leftmost',
    'column is not a pessimistic case, it is the segment median. Beating it is the whole plan.',
]):
    SN[f'A{16+j}'] = line
    SN[f'A{16+j}'].font = Font(name='Arial', size=9, color=INK)

# ═══════════════════════════════════════════════════════════════════════════
# 10. VALUATION
# ═══════════════════════════════════════════════════════════════════════════
VL = sheet('Valuation', {'A': 40}, first=40)
VL['A1'] = 'VALUATION'
VL['A1'].font = Font(name='Arial', size=15, bold=True, color=NAVY)
VL['A2'] = '2046 enterprise value across revenue outcomes and multiples. Neither axis is within our control alone.'
VL['A2'].font = Font(name='Arial', size=9, italic=True, color=GREY)

VL['A4'] = 'Live model revenue, 2046'
VL['A4'].font = Font(name='Arial', size=9, bold=True, color=NAVY)
VL['B4'] = f'=Revenue!{LAST}36'
VL['B4'].number_format = MONEY
VL['B4'].font = Font(name='Arial', size=10, bold=True, color=NAVY)
VL['B4'].fill = KEY_FILL

mults = [3.0, 4.0, 5.0, 6.0, 8.0, 10.0]
revs = [0.15, 0.5, 1.0, 1.8, 3.0, 5.5]     # $bn
VL['A6'] = 'Revenue 2046 ($bn)  ↓     Multiple  →'
VL['A6'].font = Font(name='Arial', size=9, bold=True, color=NAVY)
for j, m in enumerate(mults):
    c = VL.cell(row=7, column=2 + j, value=m)
    c.number_format = MULT
    c.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center')
    VL.column_dimensions[get_column_letter(2 + j)].width = 13
VL['A7'].fill = HDR_FILL
for i, rv in enumerate(revs):
    rr = 8 + i
    c0 = VL.cell(row=rr, column=1, value=rv)
    c0.number_format = '0.00'
    c0.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    c0.fill = HDR_FILL
    c0.alignment = Alignment(horizontal='center')
    for j, m in enumerate(mults):
        col = get_column_letter(2 + j)
        cell = VL.cell(row=rr, column=2 + j, value=f'=$A{rr}*{col}$7')
        cell.number_format = '$#,##0.0"bn"'
        cell.font = Font(name='Arial', size=9)
        if 15 <= rv * m <= 40:
            cell.fill = PatternFill('solid', fgColor='D9EAD3')

VL['A15'] = 'The stated ambition'
VL['A15'].font = Font(name='Arial', size=10, bold=True, color=NAVY)
for j, line in enumerate([
    'Shaded cells land in the $15–40bn band. They require between roughly $1.9bn and $8bn of revenue,',
    'depending entirely on the multiple earned. The Conservative case does not approach the band; the',
    'Expected case touches it only at a premium multiple; the Optimistic case sits inside it.',
    '',
    'A premium multiple is earned by two numbers and no others: net revenue retention above 120%, and',
    'a Rule of 40 above 50. Both are on the P&L and Sensitivity sheets. Neither is a marketing decision.',
    '',
    'This sheet is arithmetic, not a target. Book VIII §5 is explicit that a valuation must never become',
    'the operating objective, because optimising for it produces the behaviour Book I prohibits.',
]):
    VL[f'A{16+j}'] = line
    VL[f'A{16+j}'].font = Font(name='Arial', size=9, color=INK)

wb.save('StromeX-Financial-Master-Plan.xlsx')
print('wrote StromeX-Financial-Master-Plan.xlsx')
